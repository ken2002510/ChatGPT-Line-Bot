import openpyxl
def find_next_row_by_first_row_value(file_path, search_str):
    try:
        # 開啟 Excel 檔案
        workbook = openpyxl.load_workbook(file_path)

        # 選擇第一個工作表
        sheet = workbook.active

        # 搜尋搜尋字串所在的儲存格
        for column in sheet.iter_rows(min_row=1, max_row=sheet.max_column, min_col=1, max_col=1):
            print(column)
            for cell in column:
                if cell.value == search_str:
                    # 找到該儲存格，輸出下一行的值
                    return sheet.cell(row=cell.row , column=cell.column+1).value

        # 若搜尋字串未找到，輸出 "NO FOUND"
        return "NO 3FOUND"
    except:
        # 若有錯誤發生，輸出 "NO FOUND"
        return "NO 2FOUND"