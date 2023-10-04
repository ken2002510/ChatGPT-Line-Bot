from openpyxl import load_workbook
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name('lineBotContent.json', scope)
client = gspread.authorize(credentials)
def append_to_excel(text):
    file_path="check.xlsx"
    # 打開 Excel 文件
    wb = load_workbook(filename=file_path)

    # 選擇第一個工作表
    ws = wb.active
    
    # 找到最後一行的行數
    last_row = ws.max_row

    # 在下一行添加數據
    text="{"+text+"}"
    data = json.loads(text)  
    print(data)
    for colum, item in enumerate(data, start=1):
      ws.cell(row=last_row+1, column=colum,value = data[item])


    # 保存文件
    wb.save(file_path)

def write_to_google_sheets(data):
    try:
        sheet_name = "content"  # 將 "Your Sheet Name" 替換為你的 Google Sheets 表名稱
        sheet = client.open(sheet_name).sheet1  # 開啟表格
        row = [data]  # 要寫入的數據，這裡假設 data 是要寫入的內容
        sheet.append_row(row)  # 將數據寫入表格
        return "寫入成功"
    except Exception as e:
        return f"寫入失敗：{str(e)}"
def write_report(data):
    try:
        data="{"+data+"}"
        sheet_name = "report"  # 將 "Your Sheet Name" 替換為你的 Google Sheets 表名稱
        sheet = client.open(sheet_name).sheet1  # 開啟表格
        data_dict = json.loads(data)
        data_list = list(data_dict.values())
        sheet.append_rows([data_list])  # 将数据附加到表格
        return "寫入成功"
    except Exception as e:
        return f"寫入失敗：{str(e)}"
def read_to_google_sheets(data):
    # 打開 Google Sheets 表單
  sheet = client.open('teacher').sheet1

  # 讀取數據
  data = sheet.get_all_records()
  
  # 打印數據
  for record in data:
      print(record)
def write_content(name,data):
    # 打開 Google Sheets 表單
  sheet = client.open('content').sheet1
  cell = sheet.find(name)
  # 讀取數據
  row = cell.row
  col = cell.col
 # 讀取整行數據
  row_data = sheet.row_values(row)
  
  # 找到第一個空白的單元格，即該行的最後一列
  last_col = len(row_data) + 1
  
  # 將數據寫入最後一列
  sheet.update_cell(row, last_col, data)
  return "數據已寫入"
def read_practice(i):
    # 打開 Google Sheets 表單
  sheet = client.open('practice').sheet1
  # data_in_user_col = ["",""]
  data_in_user_col = sheet.row_values(i)
  print(data_in_user_col)
  return data_in_user_col[0],data_in_user_col[1]